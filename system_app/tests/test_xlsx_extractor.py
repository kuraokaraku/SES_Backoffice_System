"""Excelパーサのユニットテスト"""
import tempfile
import unittest

import openpyxl

from system_app.services.contract_parsers.xlsx_extractor import (
    cells_to_text,
    extract_cells,
    render_html,
)


def _make_xlsx(data: dict, merges: list[tuple] = None) -> str:
    """辞書データからテスト用Excelファイルを生成してパスを返す。"""
    wb = openpyxl.Workbook()
    ws = wb.active
    for coord, value in data.items():
        ws[coord] = value
    if merges:
        for merge_range in merges:
            ws.merge_cells(merge_range)
    tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
    wb.save(tmp.name)
    tmp.close()
    return tmp.name


class ExtractCellsTest(unittest.TestCase):
    def test_returns_non_empty_cells(self):
        path = _make_xlsx({"A1": "テスト", "B2": 123})
        cells = extract_cells(path)
        coords = {c["coord"] for c in cells}
        self.assertIn("A1", coords)
        self.assertIn("B2", coords)

    def test_empty_cells_excluded(self):
        path = _make_xlsx({"A1": "値あり"})
        cells = extract_cells(path)
        self.assertEqual(len(cells), 1)
        self.assertEqual(cells[0]["value"], "値あり")

    def test_cell_dict_has_required_keys(self):
        path = _make_xlsx({"C3": "確認"})
        cell = extract_cells(path)[0]
        for key in ("coord", "row", "col", "value"):
            self.assertIn(key, cell)

    def test_numeric_value_converted_to_string(self):
        path = _make_xlsx({"A1": 550000})
        cells = extract_cells(path)
        self.assertEqual(cells[0]["value"], "550000")


class CellsToTextTest(unittest.TestCase):
    def test_format(self):
        cells = [
            {"coord": "A1", "value": "案件名"},
            {"coord": "D1", "value": "テスト"},
        ]
        text = cells_to_text(cells)
        self.assertIn("A1: 案件名", text)
        self.assertIn("D1: テスト", text)

    def test_newline_separated(self):
        cells = [{"coord": "A1", "value": "A"}, {"coord": "A2", "value": "B"}]
        lines = cells_to_text(cells).splitlines()
        self.assertEqual(len(lines), 2)


class RenderHtmlTest(unittest.TestCase):
    def test_returns_html_table(self):
        path = _make_xlsx({"A10": "案件名", "D10": "システム開発"})
        html = render_html(path)
        self.assertIn("<table", html)
        self.assertIn("案件名", html)

    def test_highlight_coord_applies_yellow_bg(self):
        path = _make_xlsx({"A10": "単価", "D10": 500000})
        html = render_html(path, highlight_coords=["D10"])
        self.assertIn("ffe066", html)

    def test_non_highlight_does_not_have_yellow_bg(self):
        path = _make_xlsx({"A10": "単価", "D10": 500000})
        html = render_html(path, highlight_coords=[])
        self.assertNotIn("ffe066", html)

    def test_header_only_columns_excluded(self):
        # A1-A9 は本文行なし → 除外、A10 は本文行あり → 表示
        path = _make_xlsx({"A1": "ヘッダー専用", "B10": "本文データ"})
        html = render_html(path)
        # A列はヘッダー行のみなので列ヘッダーに出ないはず
        import re
        col_headers = re.findall(r"<th[^>]*>([A-Z]+)</th>", html)
        self.assertNotIn("A", col_headers)
        self.assertIn("B", col_headers)

    def test_merged_cell_origin_included(self):
        path = _make_xlsx(
            {"A10": "ラベル", "D10": "値"},
            merges=["D10:F10"],
        )
        html = render_html(path)
        self.assertIn("値", html)
