"""請求書Excelエクスポート（雛型テンプレートへのセル埋め込み）"""
import os
from datetime import date

import openpyxl
from django.conf import settings

from system_app.models import Invoice
from system_app.services.contracts import get_active_contract
from system_app.services.invoice_calculator import default_due_date

TEMPLATE_PATH = os.path.join(
    settings.BASE_DIR, "system_app", "templates", "【研修用】【雛型】請求書_SES_ITFL.xlsx"
)

WEEKDAY_JA = ["月", "火", "水", "木", "金", "土", "日"]


def export_invoice_to_template_xlsx(invoice_id, template_path=None):
    """
    Invoice からテンプレートExcelにセルを埋めて保存する。

    Returns
    -------
    dict  {"file_path": str, "file_name": str, "content_type": str}
    """
    template_path = template_path or TEMPLATE_PATH

    invoice = (
        Invoice.objects
        .select_related("assignment", "assignment__upstream_entity")
        .get(id=invoice_id)
    )
    lines = {line.kind: line for line in invoice.lines.all()}

    wb = openpyxl.load_workbook(template_path)
    ws = wb["【雛型】請求書"]

    # --- ヘッダ ---
    ws["A4"] = invoice.assignment.upstream_entity.name
    ws["H2"] = invoice.invoice_number or f"DRAFT-{invoice.id}"

    issue = invoice.issue_date or date.today()
    ws["F5"] = issue

    # --- 支払期日（必ず上書きして雛型の古い日付を潰す） ---
    due = invoice.due_date
    if not due:
        try:
            contract = get_active_contract(invoice.assignment, invoice.billing_ym)
            terms = contract.upstream_payment_terms
        except Exception:
            terms = None
        due = default_due_date(invoice.billing_ym, terms)
    ws["A12"] = due
    ws["B12"] = f"（{WEEKDAY_JA[due.weekday()]}）"

    # --- 明細: 基本行 ---
    basic = lines.get("basic")
    if basic:
        ws["A16"] = basic.item_name
        ws["C16"] = float(basic.quantity)
        ws["E16"] = float(basic.unit_price)
        ws["G16"] = float(basic.amount)

    # 月分表示
    month = int(invoice.billing_ym[4:])
    ws["A17"] = f"{month}月分"

    # --- 明細: 超過（無い場合も0で埋めて式の破綻を防ぐ） ---
    excess = lines.get("excess")
    if excess:
        ws["C18"] = float(excess.quantity)
        ws["E18"] = float(excess.unit_price)
        ws["G18"] = float(excess.amount)
    else:
        ws["C18"] = 0
        ws["E18"] = 0
        ws["G18"] = 0

    # --- 明細: 控除（無い場合も0で埋める） ---
    deduction = lines.get("deduction")
    if deduction:
        ws["C19"] = float(deduction.quantity)
        ws["E19"] = float(deduction.unit_price)
        ws["G19"] = float(deduction.amount)
    else:
        ws["C19"] = 0
        ws["E19"] = 0
        ws["G19"] = 0

    # --- 集計（値で上書き） ---
    ws["G20"] = float(invoice.subtotal_amount)
    ws["G21"] = float(invoice.tax_amount)
    ws["G22"] = float(invoice.subtotal_amount + invoice.tax_amount)

    # --- 交通費 ---
    expense = lines.get("expense")
    ws["G25"] = float(expense.amount) if expense else 0

    # --- 合計 ---
    ws["G27"] = float(invoice.total_amount)

    # --- 金額ヘッダ ---
    ws["E14"] = float(invoice.total_amount)
    ws["H14"] = float(invoice.tax_amount)

    # --- 保存 ---
    invoice_label = invoice.invoice_number or str(invoice.id)
    output_dir = os.path.join(
        settings.MEDIA_ROOT, "invoices", invoice.billing_ym, invoice_label
    )
    os.makedirs(output_dir, exist_ok=True)

    file_name = f"請求書_{invoice.billing_ym}_{invoice_label}.xlsx"
    file_path = os.path.join(output_dir, file_name)
    wb.save(file_path)
    wb.close()

    return {
        "file_path": file_path,
        "file_name": file_name,
        "content_type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    }
