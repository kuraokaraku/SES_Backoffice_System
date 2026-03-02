"""見積書Excelエクスポート（雛型テンプレートへのセル埋め込み）"""
import os
import tempfile
from datetime import date

import openpyxl
from django.conf import settings

from system_app.models import Assignment, ServiceContract

TEMPLATE_PATH = os.path.join(
    settings.BASE_DIR, "system_app", "templates", "【雛型】御見積書_ITFL.xlsx"
)


def export_estimate_xlsx(assignment_id, contract_id=None, template_path=None):
    """
    Assignment + ServiceContract から雛型Excelにセルを埋めて保存する。

    Returns
    -------
    dict  {"file_path": str, "file_name": str, "content_type": str}
    """
    template_path = template_path or TEMPLATE_PATH

    assignment = (
        Assignment.objects
        .select_related(
            "worker_entity",
            "upstream_entity",
            "upstream_contact_person",
        )
        .get(id=assignment_id)
    )

    # 契約取得: 指定IDがあればそれ、なければ現行有効な最新契約
    if contract_id:
        contract = ServiceContract.objects.get(id=contract_id, assignment=assignment)
    else:
        contract = _get_active_contract(assignment)

    wb = openpyxl.load_workbook(template_path)
    ws = wb["雛型"]

    today = date.today()

    # --- ヘッダ ---
    ws["G2"] = f"{today.year}年{today.month}月{today.day}日"

    # 宛先
    client_name = assignment.upstream_entity.name if assignment.upstream_entity else ""
    ws["A5"] = f"{client_name}　　御中"

    # A12: 甲乙表記（クライアント名を差し込み）
    ws["A12"] = f"（以下、発注者・{client_name}を「甲」、受託者・株式会社ITFLを「乙」と呼称する。）"

    # ご担当
    contact_name = ""
    if assignment.upstream_contact_person:
        contact_name = assignment.upstream_contact_person.name
    ws["A7"] = f"ご担当：　{contact_name}　様" if contact_name else ""

    # --- 契約条件 ---
    # 業務名
    ws["C13"] = assignment.project_name or ""

    # 契約形態
    ws["C14"] = "SES（準委任契約）"

    # 作業期間
    if contract:
        from_str = contract.valid_from.strftime("%Y/%m/%d") if contract.valid_from else "●●●●/●●/●●"
        to_str = contract.valid_to.strftime("%Y/%m/%d") if contract.valid_to else "●●●●/●●/●●"
        ws["C15"] = f"{from_str}　～　{to_str}"

    # 作業者名
    worker_name = assignment.worker_entity.name if assignment.worker_entity else ""
    ws["C19"] = worker_name

    if contract:
        # 金額（税抜）
        ws["D20"] = f"¥{contract.unit_price:,}/月" if contract.unit_price else ""

        # 精算
        if contract.lower_limit_hour is not None and contract.upper_limit_hours is not None:
            lower = _format_hours(contract.lower_limit_hour)
            upper = _format_hours(contract.upper_limit_hours)
            ws["C21"] = f"有（{lower}h-{upper}h）"
        elif contract.is_fixed_fee:
            ws["C21"] = "無（固定報酬）"
        else:
            ws["C21"] = ""

        # 精算単位
        if contract.settlement_unit_minutes:
            ws["C22"] = f"{contract.settlement_unit_minutes}分"

        # 超過単価
        if contract.excess_unit_price:
            ratio = _calc_ratio(contract.excess_unit_price, contract.unit_price)
            ws["C23"] = f"¥{contract.excess_unit_price:,}/h　({ratio})"

        # 控除単価
        if contract.deduction_unit_price:
            ratio = _calc_ratio(contract.deduction_unit_price, contract.unit_price)
            ws["C24"] = f"¥{contract.deduction_unit_price:,}/h　({ratio})"

        # 支払条件
        if contract.upstream_payment_terms:
            days = contract.upstream_payment_terms
            holiday = contract.bank_holiday_handling or "翌営業日払い"
            # 支払サイト → 翌月/翌々月 + 日付に変換
            if days <= 30:
                month_label = "翌月"
                pay_day = days
            else:
                month_label = "翌々月"
                pay_day = days - 30
            if pay_day == 30:
                day_str = "末日"
            else:
                day_str = f"{pay_day}日"
            ws["C25"] = f" 月末〆{month_label}{day_str}支払い（{days}日サイト（金融機関休業日の場合は{holiday}））"

    # --- 保存 ---
    out_dir = tempfile.mkdtemp()
    file_name = f"御見積書_{client_name}_{worker_name}_{today.strftime('%Y%m%d')}.xlsx"
    file_path = os.path.join(out_dir, file_name)
    wb.save(file_path)

    return {
        "file_path": file_path,
        "file_name": file_name,
        "content_type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    }


def _get_active_contract(assignment):
    """現行有効な契約を取得（valid_to が NULL or 今日以降の最新）"""
    from django.db.models import Q
    today = date.today()
    return (
        assignment.contracts
        .filter(Q(valid_to__isnull=True) | Q(valid_to__gte=today))
        .order_by("-valid_from")
        .first()
    )


def _format_hours(val):
    """Decimal を整数 or 小数1桁で表示"""
    if val is None:
        return "●"
    if val == int(val):
        return str(int(val))
    return f"{val:.1f}"


def _calc_ratio(part_price, total_price):
    """単価の割合を計算して表示（例: 5割, 7割）"""
    if not total_price or not part_price:
        return "●割"
    # 時間単価 / (月額 / 精算中央) は複雑なので、
    # ここではシンプルに「●割」として返す（実運用で調整）
    return "●割"
