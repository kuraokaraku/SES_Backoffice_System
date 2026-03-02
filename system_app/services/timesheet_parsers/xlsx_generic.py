"""汎用 Excel 勤務表パーサ（キーワードスコアリング方式）"""
import re
from datetime import date, datetime, timedelta
from decimal import Decimal, InvalidOperation
from pathlib import Path

import openpyxl

# =====================================================
# 定数
# =====================================================

SHEET_PREFER = ["勤務", "作業", "稼働", "実績", "勤務表"]
SHEET_PENALISE = ["記入方法", "説明", "祝日", "マスタ", "データ"]

BYM_KEYWORDS = ["対象年月", "請求月", "作業月", "勤務月", "対象月", "稼働月", "月度", "年月"]

HOURS_STRONG = [
    "実働時間合計", "就業時間合計", "稼働時間合計",
    "勤務時間合計", "作業時間合計", "合計時間", "月間合計",
]
HOURS_WEAK = ["実働", "就業", "稼働", "勤務", "作業", "時間"]
SUM_WORDS = ["合計", "総計", "月計", "計", "TOTAL", "Total"]

TRAVEL_KEYWORDS = ["交通費", "通勤費", "旅費", "経費", "立替", "立替金", "実費"]

# セル探索の最大距離
MAX_COL_DISTANCE = 12
MAX_ROW_DISTANCE = 3

# =====================================================
# ユーティリティ
# =====================================================

def _norm_text(value):
    """セル値を正規化テキストに変換。"""
    if value is None:
        return ""
    s = str(value).strip()
    s = s.replace("\u3000", " ")  # 全角空白
    s = re.sub(r"\n", " ", s)
    s = re.sub(r"\s+", " ", s)
    return s


def _to_decimal(value):
    """セル値を Decimal に変換できれば返す。"""
    if value is None:
        return None
    if isinstance(value, timedelta):
        return Decimal(str(round(value.total_seconds() / 3600, 2)))
    if isinstance(value, (int, float)):
        return Decimal(str(value))
    if isinstance(value, str):
        # "hh:mm" 形式
        m = re.match(r"^(\d{1,4}):(\d{2})$", value.strip())
        if m:
            return Decimal(str(int(m.group(1)) + int(m.group(2)) / 60))
        # 純粋な数値文字列
        try:
            return Decimal(value.strip().replace(",", ""))
        except InvalidOperation:
            pass
    return None


def _extract_hours_from_text(text):
    """文字列に埋め込まれた時間を抽出。例: '就業時間合計 185:30 H'"""
    # hh:mm パターン
    m = re.search(r"(\d{1,4}):(\d{2})", text)
    if m:
        return Decimal(str(int(m.group(1)) + int(m.group(2)) / 60))
    # 小数パターン
    m = re.search(r"(\d{1,4}(?:\.\d+))", text)
    if m:
        return Decimal(m.group(1))
    return None


def _extract_amount_from_text(text):
    """文字列に埋め込まれた金額を抽出。"""
    m = re.search(r"[\d,]+", text.replace(" ", ""))
    if m:
        try:
            return Decimal(m.group(0).replace(",", ""))
        except InvalidOperation:
            pass
    return None


def _sheet_score(sheet_name):
    """シート名による加点/減点。"""
    score = 0.0
    for kw in SHEET_PREFER:
        if kw in sheet_name:
            score += 0.10
            break
    for kw in SHEET_PENALISE:
        if kw in sheet_name:
            score -= 0.10
            break
    return score


def _clamp(val, lo=0.0, hi=1.0):
    return max(lo, min(hi, val))


def _cell_ref(row, col):
    """(row, col) を "A1" 形式に変換。"""
    from openpyxl.utils import get_column_letter
    return f"{get_column_letter(col)}{row}"


def _contains_any(text, keywords):
    """text がキーワードのいずれかを含むか。"""
    for kw in keywords:
        if kw in text:
            return True
    return False


def _nearby_text(ws, row, col, distance=2):
    """(row, col) の近傍セルのテキストを結合して返す。"""
    texts = []
    for dr in range(-distance, distance + 1):
        for dc in range(-distance, distance + 1):
            if dr == 0 and dc == 0:
                continue
            r, c = row + dr, col + dc
            if r < 1 or c < 1:
                continue
            try:
                v = ws.cell(row=r, column=c).value
                if v is not None:
                    texts.append(_norm_text(v))
            except Exception:
                pass
    return " ".join(texts)


def _row_context(ws, row, max_col=None):
    """指定行の全セルを結合してコンテキスト文字列を返す。"""
    if max_col is None:
        max_col = ws.max_column or 1
    parts = []
    for c in range(1, max_col + 1):
        try:
            v = ws.cell(row=row, column=c).value
            if v is not None:
                parts.append(_norm_text(v))
        except Exception:
            pass
    return " | ".join(parts) if parts else ""


def _build_cell_context(ws, row, col):
    """候補セルの周辺コンテキストを構築する（同じ行 + 1行上）。最大150文字。"""
    max_col = ws.max_column or col
    same_row = _row_context(ws, row, max_col)
    above = _row_context(ws, row - 1, max_col) if row > 1 else ""
    if above:
        ctx = f"[上行] {above} [当行] {same_row}"
    else:
        ctx = same_row
    return ctx[:150]


def _search_value_right_down(ws, row, col, max_col=MAX_COL_DISTANCE, max_row=MAX_ROW_DISTANCE):
    """アンカーセルから右方向→下方向に数値セルを探す。"""
    results = []
    # 右方向
    for dc in range(1, max_col + 1):
        c = col + dc
        try:
            cell = ws.cell(row=row, column=c)
        except Exception:
            break
        val = _to_decimal(cell.value)
        if val is not None:
            results.append({
                "value": val,
                "row": row,
                "col": c,
                "direction": "right",
                "distance": dc,
            })
    # 下方向
    for dr in range(1, max_row + 1):
        r = row + dr
        try:
            cell = ws.cell(row=r, column=col)
        except Exception:
            break
        val = _to_decimal(cell.value)
        if val is not None:
            results.append({
                "value": val,
                "row": r,
                "col": col,
                "direction": "down",
                "distance": dr,
            })
    return results


# =====================================================
# A. billing_ym 抽出
# =====================================================

_RE_YM_SLASH = re.compile(r"(20\d{2})[/\-\.](0?[1-9]|1[0-2])(?![0-9])")
_RE_YM_JA = re.compile(r"(20\d{2})\s*年\s*(0?[1-9]|1[0-2])\s*月")
_RE_YM_COMPACT = re.compile(r"^(20\d{2})(0[1-9]|1[0-2])$")


def _extract_ym_from_text(text):
    """1セルからYYYYMMを抽出。"""
    for pat in [_RE_YM_JA, _RE_YM_SLASH, _RE_YM_COMPACT]:
        m = pat.search(text)
        if m:
            return f"{m.group(1)}{int(m.group(2)):02d}"
    return None


def _find_billing_ym_candidates(ws, sheet_name):
    """billing_ym の候補リストを返す。"""
    candidates = []
    s_score = _sheet_score(sheet_name)

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column):
        for cell in row:
            text = _norm_text(cell.value)
            if not text:
                continue
            r, c = cell.row, cell.column

            # ルート1: 1セルに年月が入っている
            ym = _extract_ym_from_text(text)
            if ym:
                score = 0.55
                nearby = _nearby_text(ws, r, c)
                if _contains_any(text + " " + nearby, BYM_KEYWORDS):
                    score += 0.25
                if r <= 6:
                    score += 0.20
                if r >= 20:
                    score -= 0.10
                # 説明文の誤検出を減点
                if _contains_any(text, ["等", "例", "入力", "※", "記入"]):
                    score -= 0.30
                score += s_score
                candidates.append({
                    "value": ym,
                    "cell": _cell_ref(r, c),
                    "sheet": sheet_name,
                    "confidence": _clamp(score),
                    "evidence": f"セル{_cell_ref(r, c)}から年月パターン検出: {text[:40]}",
                })
                continue

            # ルート2: 年セル+月セルが別
            if isinstance(cell.value, (int, float)) and 2020 <= cell.value <= 2099:
                year = int(cell.value)
                # 同一行で "月" の近くに 1-12 を探す
                for dc in range(1, 8):
                    nc = c + dc
                    try:
                        nval = ws.cell(row=r, column=nc).value
                    except Exception:
                        break
                    if isinstance(nval, (int, float)) and 1 <= nval <= 12:
                        ym = f"{year}{int(nval):02d}"
                        score = 0.55
                        row_text = " ".join(
                            _norm_text(ws.cell(row=r, column=cc).value)
                            for cc in range(max(1, c - 2), min((ws.max_column or c) + 1, nc + 3))
                        )
                        if "年" in row_text and ("月" in row_text or "月度" in row_text):
                            score += 0.25
                        if r <= 6:
                            score += 0.20
                        if r >= 20:
                            score -= 0.10
                        score += s_score
                        candidates.append({
                            "value": ym,
                            "cell": f"{_cell_ref(r, c)}+{_cell_ref(r, nc)}",
                            "sheet": sheet_name,
                            "confidence": _clamp(score),
                            "evidence": f"年={year}({_cell_ref(r, c)}), 月={int(nval)}({_cell_ref(r, nc)})",
                        })
                        break

            # ルート3: date/datetime セルから推定
            if isinstance(cell.value, (date, datetime)):
                dt = cell.value if isinstance(cell.value, date) else cell.value.date()
                ym = f"{dt.year}{dt.month:02d}"
                score = 0.40
                nearby = _nearby_text(ws, r, c)
                if _contains_any(text + " " + nearby, BYM_KEYWORDS):
                    score += 0.25
                if r <= 6:
                    score += 0.20
                if r >= 20:
                    score -= 0.10
                score += s_score
                candidates.append({
                    "value": ym,
                    "cell": _cell_ref(r, c),
                    "sheet": sheet_name,
                    "confidence": _clamp(score),
                    "evidence": f"日付セル{_cell_ref(r, c)}: {dt.isoformat()}",
                })

    return candidates


# =====================================================
# B. actual_hours 抽出
# =====================================================

def _find_actual_hours_candidates(ws, sheet_name):
    """actual_hours の候補リストを返す。"""
    candidates = []
    s_score = _sheet_score(sheet_name)

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column):
        for cell in row:
            text = _norm_text(cell.value)
            if not text:
                continue
            r, c = cell.row, cell.column

            # 強キーワードチェック
            strong_match = any(kw in text for kw in HOURS_STRONG)
            # 弱キーワード + 合計語（セル内 or 近傍）
            nearby = _nearby_text(ws, r, c)
            combined = text + " " + nearby + " " + sheet_name
            weak_match = (
                any(kw in combined for kw in HOURS_WEAK)
                and any(sw in text for sw in SUM_WORDS)
            )
            # 「合計」単体セル + 近傍/シート名に時間系キーワードがあるケース
            sum_only_match = (
                any(sw == text for sw in SUM_WORDS)
                and any(kw in nearby + " " + sheet_name for kw in HOURS_WEAK + ["H", "ｈ", "時間"])
            )

            if not strong_match and not weak_match and not sum_only_match:
                continue

            base_score = 0.45 if strong_match else (0.30 if weak_match else 0.25)
            if _contains_any(nearby, SUM_WORDS):
                base_score += 0.20

            # ルート2: 値がアンカーセル自体に埋め込まれている
            embedded = _extract_hours_from_text(text)
            if embedded is not None and 0 <= embedded <= 400:
                score = base_score + 0.20
                if r >= 20:
                    score += 0.10
                if 60 <= embedded <= 260:
                    score += 0.10
                score += s_score
                candidates.append({
                    "value": embedded,
                    "cell": _cell_ref(r, c),
                    "sheet": sheet_name,
                    "confidence": _clamp(score),
                    "evidence": f"セル{_cell_ref(r, c)}に埋め込み値: {text[:50]}",
                    "context": _build_cell_context(ws, r, c),
                })

            # ルート1: 右方向/下方向探索
            found = _search_value_right_down(ws, r, c)
            for f in found:
                val = f["value"]
                if val < 0 or val > 400:
                    continue
                score = base_score
                if f["direction"] == "right":
                    score += 0.20 * max(0, 1 - f["distance"] / MAX_COL_DISTANCE)
                if r >= 20:
                    score += 0.10
                score += 0.20  # 数値変換成功
                if 60 <= val <= 260:
                    score += 0.10
                if 1 <= val <= 31:
                    score -= 0.15  # 日付っぽい
                if val == 0:
                    score -= 0.10
                score += s_score
                candidates.append({
                    "value": val,
                    "cell": _cell_ref(f["row"], f["col"]),
                    "sheet": sheet_name,
                    "confidence": _clamp(score),
                    "evidence": f"{_cell_ref(r, c)}に'{text[:30]}', {f['direction']}方向{_cell_ref(f['row'], f['col'])}に数値{val}",
                    "context": _build_cell_context(ws, f["row"], f["col"]),
                })

    return candidates


# =====================================================
# C. travel_amount 抽出
# =====================================================

def _find_travel_amount_candidates(ws, sheet_name):
    """travel_amount の候補リストを返す。"""
    candidates = []
    s_score = _sheet_score(sheet_name)

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column):
        for cell in row:
            text = _norm_text(cell.value)
            if not text:
                continue
            r, c = cell.row, cell.column

            if not _contains_any(text, TRAVEL_KEYWORDS):
                continue

            base_score = 0.45
            nearby = _nearby_text(ws, r, c)
            if _contains_any(nearby, SUM_WORDS):
                base_score += 0.15

            # 埋め込み金額
            embedded = _extract_amount_from_text(text)
            if embedded is not None and 0 < embedded <= 500000:
                score = base_score + 0.20
                score += s_score
                candidates.append({
                    "value": embedded,
                    "cell": _cell_ref(r, c),
                    "sheet": sheet_name,
                    "confidence": _clamp(score),
                    "evidence": f"セル{_cell_ref(r, c)}に埋め込み金額: {text[:50]}",
                })

            # 右方向/下方向探索
            found = _search_value_right_down(ws, r, c)
            for f in found:
                val = f["value"]
                if val <= 0 or val > 500000:
                    continue
                score = base_score
                if f["direction"] == "right":
                    score += 0.20 * max(0, 1 - f["distance"] / MAX_COL_DISTANCE)
                score += 0.20  # 金額として妥当
                if val == 0:
                    score -= 0.10
                score += s_score
                candidates.append({
                    "value": val,
                    "cell": _cell_ref(f["row"], f["col"]),
                    "sheet": sheet_name,
                    "confidence": _clamp(score),
                    "evidence": f"{_cell_ref(r, c)}に'{text[:30]}', {f['direction']}方向{_cell_ref(f['row'], f['col'])}に金額{val}",
                })

    return candidates


# =====================================================
# メインエントリ
# =====================================================

def _pick_best(candidates, threshold):
    """候補リストからベスト1を選び、閾値判定する。"""
    if not candidates:
        return None, candidates
    sorted_c = sorted(candidates, key=lambda x: x["confidence"], reverse=True)
    best = sorted_c[0]
    # 曖昧判定: 1位と2位の差が小さい
    if len(sorted_c) >= 2 and str(best["value"]) != str(sorted_c[1]["value"]) and (best["confidence"] - sorted_c[1]["confidence"]) < 0.05:
        return None, sorted_c[:5]
    if best["confidence"] >= threshold:
        return best, sorted_c[:5]
    return None, sorted_c[:5]


def parse_timesheet_xlsx_generic(path):
    """
    汎用 Excel 勤務表パーサ。

    Parameters
    ----------
    path : str | Path
        xlsx ファイルパス

    Returns
    -------
    dict  仕様書記載の出力形式
    """
    wb = openpyxl.load_workbook(str(path), data_only=True)

    all_bym = []
    all_hours = []
    all_travel = []

    for ws in wb.worksheets:
        sn = ws.title
        all_bym.extend(_find_billing_ym_candidates(ws, sn))
        all_hours.extend(_find_actual_hours_candidates(ws, sn))
        all_travel.extend(_find_travel_amount_candidates(ws, sn))

    wb.close()

    # 重複除去（同一セル・同一値）
    def _dedup(cands):
        seen = set()
        out = []
        for c in cands:
            key = (c["cell"], str(c["value"]), c["sheet"])
            if key not in seen:
                seen.add(key)
                out.append(c)
        return out

    all_bym = _dedup(all_bym)
    all_hours = _dedup(all_hours)
    all_travel = _dedup(all_travel)

    best_bym, cands_bym = _pick_best(all_bym, threshold=0.75)
    best_hours, cands_hours = _pick_best(all_hours, threshold=0.70)
    best_travel, cands_travel = _pick_best(all_travel, threshold=0.65)

    def _to_result(best):
        if best is None:
            return {"value": None, "confidence": 0.0, "cell": None, "sheet": None, "evidence": ""}
        return {
            "value": best["value"],
            "confidence": best["confidence"],
            "cell": best["cell"],
            "sheet": best.get("sheet"),
            "evidence": best["evidence"],
        }

    return {
        "billing_ym": _to_result(best_bym),
        "actual_hours": _to_result(best_hours),
        "travel_amount": _to_result(best_travel),
        "candidates": {
            "billing_ym": cands_bym,
            "actual_hours": cands_hours,
            "travel_amount": cands_travel,
        },
        "parser": "generic_keyword_scoring_v2",
    }
