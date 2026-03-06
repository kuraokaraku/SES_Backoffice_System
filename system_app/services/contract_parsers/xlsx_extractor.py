"""Excel 契約書パーサ: セル値抽出 + HTML変換 + ハイライト"""
import re
import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string


def extract_cells(file_path):
    """
    Excelから全セル値を取得する。

    Returns
    -------
    list[dict]  {"coord": "A1", "row": 1, "col": 1, "value": "..."}
    """
    wb = openpyxl.load_workbook(file_path, data_only=True)
    ws = wb.active

    cells = []
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is not None:
                cells.append({
                    "coord": cell.coordinate,
                    "row": cell.row,
                    "col": cell.column,
                    "value": str(cell.value).strip(),
                })
    return cells


def cells_to_text(cells):
    """セルリストを LLM 用のテキストに変換する。"""
    lines = [f"{c['coord']}: {c['value']}" for c in cells]
    return "\n".join(lines)


def _fmt_value(val):
    """セル値を表示用文字列に変換する。日時は日付のみに整形。"""
    import datetime
    if val is None:
        return ""
    if isinstance(val, (datetime.datetime,)):
        return val.strftime("%Y-%m-%d")
    if isinstance(val, datetime.date):
        return val.strftime("%Y-%m-%d")
    return str(val).strip()


def render_html(file_path, highlight_coords=None):
    """
    Excelをシンプルな HTML テーブルに変換する。
    highlight_coords: ハイライトするセル座標のリスト (例: ["D17", "G19"])
    """
    highlight_coords = set(highlight_coords or [])
    wb = openpyxl.load_workbook(file_path, data_only=True)
    ws = wb.active

    # マージセル情報を収集
    merged = {}
    for merge in ws.merged_cells.ranges:
        min_row, min_col, max_row_m, max_col_m = (
            merge.min_row, merge.min_col, merge.max_row, merge.max_col
        )
        for r in range(min_row, max_row_m + 1):
            for c in range(min_col, max_col_m + 1):
                if r == min_row and c == min_col:
                    merged[(r, c)] = {
                        "rowspan": max_row_m - min_row + 1,
                        "colspan": max_col_m - min_col + 1,
                        "is_origin": True,
                    }
                else:
                    merged[(r, c)] = {"is_origin": False}

    # 表示する列を決定
    # 「有効列」= 値を持つセルが存在し、かつそのセルが結合起点か単独セルである列
    # さらに、ヘッダー行(1-9)にしか値がない列は除外（文書番号・日付・差出人住所など）
    HEADER_ROW_THRESHOLD = 10
    PRINT_AREA_MAX_COL = 14  # N列まで（契約条件の主要コンテンツ範囲）

    actual_max_col = 1
    valid_cols = set()
    cols_with_body_data = set()  # 本文行(10行以降)にデータがある列

    for row in ws.iter_rows():
        for cell in row:
            if cell.value is None:
                continue
            if cell.column > actual_max_col:
                actual_max_col = cell.column
            mi = merged.get((cell.row, cell.column))
            # 非起点の結合セルは無視（値は起点が代表する）
            if mi and not mi.get("is_origin"):
                continue
            valid_cols.add(cell.column)
            if cell.row >= HEADER_ROW_THRESHOLD:
                cols_with_body_data.add(cell.column)

    # ハイライト列は常に含める
    for coord in highlight_coords:
        m = re.match(r"([A-Z]+)", coord)
        if m:
            col_idx = column_index_from_string(m.group(1))
            valid_cols.add(col_idx)
            cols_with_body_data.add(col_idx)

    # ヘッダー行のみの列と印刷範囲外の列を除外
    valid_cols = {c for c in valid_cols
                  if c in cols_with_body_data and c <= PRINT_AREA_MAX_COL}

    visible_cols = sorted(c for c in valid_cols if c <= actual_max_col)

    if not visible_cols:
        visible_cols = list(range(1, actual_max_col + 1))

    # 列幅（Excelから取得、1文字≒7px）
    col_widths = {}
    for col_letter, col_dim in ws.column_dimensions.items():
        col_idx = column_index_from_string(col_letter)
        w = col_dim.width or 8
        col_widths[col_idx] = max(40, min(int(w * 7), 250))

    def colspan_for(orig_colspan, start_col):
        """元のcolspanを、表示列のみに絞った実効colspanに変換する。"""
        end_col = start_col + orig_colspan - 1
        display_cols_in_range = sum(1 for c in visible_cols if start_col <= c <= end_col)
        return max(1, display_cols_in_range)

    # 列ヘッダー行
    th_style = "background:#e9ecef; border:1px solid #ccc; padding:3px 6px; font-size:11px; text-align:center; white-space:nowrap;"
    header_tds = [f'<th style="{th_style} min-width:28px;">#</th>']
    for col_idx in visible_cols:
        w = col_widths.get(col_idx, 70)
        header_tds.append(
            f'<th style="{th_style} width:{w}px;">{get_column_letter(col_idx)}</th>'
        )
    html_rows = [f"<tr>{''.join(header_tds)}</tr>"]

    for row_idx in range(1, ws.max_row + 1):
        # 表示列に値があるかチェック
        row_vals = [ws.cell(row=row_idx, column=c).value for c in visible_cols]
        row_has_data = any(v is not None for v in row_vals)
        row_has_highlight = any(
            f"{get_column_letter(c)}{row_idx}" in highlight_coords for c in visible_cols
        )
        if not row_has_data and not row_has_highlight:
            continue

        row_bg = "#fafafa" if row_idx % 2 == 0 else "#ffffff"
        rn_style = "background:#e9ecef; border:1px solid #ccc; padding:2px 4px; font-size:10px; text-align:center; color:#888;"
        tds = [f'<td style="{rn_style}">{row_idx}</td>']

        skip_until_col = -1
        for col_idx in visible_cols:
            if col_idx <= skip_until_col:
                continue

            coord = f"{get_column_letter(col_idx)}{row_idx}"
            merge_info = merged.get((row_idx, col_idx))

            # 非起点の結合セルはスキップ
            if merge_info and not merge_info.get("is_origin"):
                continue

            cell = ws.cell(row=row_idx, column=col_idx)
            val_str = _fmt_value(cell.value)

            attrs = ""
            if merge_info and merge_info.get("is_origin"):
                rs = merge_info["rowspan"]
                cs = colspan_for(merge_info["colspan"], col_idx)
                if rs > 1:
                    attrs += f' rowspan="{rs}"'
                if cs > 1:
                    attrs += f' colspan="{cs}"'
                    skip_until_col = col_idx + merge_info["colspan"] - 1

            if coord in highlight_coords:
                bg = "#ffe066"
                border = "2px solid #f0a500"
                fw = "bold"
                title = f' title="{coord}"'
            else:
                bg = row_bg
                border = "1px solid #dee2e6"
                fw = "normal"
                title = ""

            is_highlight = coord in highlight_coords
            inner_style = (
                "max-height:3.5em; overflow:hidden;"
                if not is_highlight else ""
            )
            style = (
                f"background:{bg}; border:{border}; padding:4px 7px; "
                f"font-size:12px; white-space:pre-wrap; vertical-align:top; "
                f"font-weight:{fw}; word-break:break-word;"
            )
            inner = f'<div style="{inner_style}">{val_str}</div>' if inner_style else val_str
            tds.append(f'<td{attrs} style="{style}" data-coord="{coord}"{title}>{inner}</td>')

        html_rows.append(f"<tr>{''.join(tds)}</tr>")

    table = (
        '<div style="overflow:auto;">'
        '<table style="border-collapse:collapse; font-size:12px;">'
        + "".join(html_rows)
        + "</table></div>"
    )
    return table
