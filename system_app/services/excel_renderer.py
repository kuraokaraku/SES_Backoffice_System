"""Excel (xlsx) → HTML テーブル変換レンダラ"""
import re
from pathlib import Path

import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string


def _color_to_hex(color):
    """openpyxl Color オブジェクトを #RRGGBB に変換。"""
    if color is None:
        return None
    if color.type == "rgb" and color.rgb and color.rgb != "00000000":
        rgb = str(color.rgb)
        if len(rgb) == 8:
            return f"#{rgb[2:]}"
        if len(rgb) == 6:
            return f"#{rgb}"
    if color.type == "theme":
        return None
    return None


def _border_side_css(side):
    """openpyxl Border Side → CSS border 値。"""
    if side is None or side.style is None:
        return None
    style_map = {
        "thin": "1px solid",
        "medium": "2px solid",
        "thick": "3px solid",
        "double": "3px double",
        "dotted": "1px dotted",
        "dashed": "1px dashed",
        "hair": "1px solid",
        "mediumDashed": "2px dashed",
        "dashDot": "1px dashed",
        "mediumDashDot": "2px dashed",
        "dashDotDot": "1px dashed",
        "mediumDashDotDot": "2px dashed",
        "slantDashDot": "1px dashed",
    }
    css_style = style_map.get(side.style, "1px solid")
    color = _color_to_hex(side.color) if side.color else "#000"
    if color is None:
        color = "#000"
    return f"{css_style} {color}"


def _cell_style_css(cell):
    """セルのスタイルをインラインCSS文字列に変換。"""
    parts = []

    # フォント
    font = cell.font
    if font:
        if font.bold:
            parts.append("font-weight:bold")
        if font.size:
            parts.append(f"font-size:{font.size}pt")
        color = _color_to_hex(font.color) if font.color else None
        if color:
            parts.append(f"color:{color}")

    # 背景色
    fill = cell.fill
    if fill and fill.fgColor:
        bg = _color_to_hex(fill.fgColor)
        if bg and bg.lower() != "#000000":
            parts.append(f"background-color:{bg}")

    # 罫線
    border = cell.border
    if border:
        for side_name, css_prop in [
            ("top", "border-top"),
            ("right", "border-right"),
            ("bottom", "border-bottom"),
            ("left", "border-left"),
        ]:
            side = getattr(border, side_name, None)
            val = _border_side_css(side)
            if val:
                parts.append(f"{css_prop}:{val}")

    # 配置
    alignment = cell.alignment
    if alignment:
        if alignment.horizontal:
            ha_map = {"left": "left", "center": "center", "right": "right",
                      "general": "left", "justify": "justify"}
            ha = ha_map.get(alignment.horizontal)
            if ha:
                parts.append(f"text-align:{ha}")
        if alignment.vertical:
            va_map = {"top": "top", "center": "middle", "bottom": "bottom"}
            va = va_map.get(alignment.vertical)
            if va:
                parts.append(f"vertical-align:{va}")

    return ";".join(parts)


def _format_cell_value(cell):
    """セル値を表示用文字列に変換。"""
    val = cell.value
    if val is None:
        return ""
    if isinstance(val, float):
        if val == int(val):
            return str(int(val))
        return str(val)
    return str(val)


def _get_data_bounds(ws, max_row_limit=500, empty_row_cutoff=5):
    """実データ範囲を特定し、空行連続で打ち切り。"""
    max_col = ws.max_column or 1
    effective_max_row = min(ws.max_row or 1, max_row_limit)

    consecutive_empty = 0
    last_data_row = 1

    for r in range(1, effective_max_row + 1):
        has_data = False
        for c in range(1, max_col + 1):
            if ws.cell(row=r, column=c).value is not None:
                has_data = True
                break
        if has_data:
            consecutive_empty = 0
            last_data_row = r
        else:
            consecutive_empty += 1
            if consecutive_empty >= empty_row_cutoff:
                break

    return last_data_row, max_col


def render_excel_to_html(file_path, highlight_cells=None):
    """
    Excelファイルを読み込み、シート別のHTML文字列を返す。

    Parameters
    ----------
    file_path : str
        xlsxファイルパス
    highlight_cells : dict | None
        {"sheet_name": ["A3", "F12", ...]} 形式。該当セルに .highlight クラスを付与

    Returns
    -------
    list[dict]
        [{"name": "シート名", "html": "<table>...</table>"}, ...]
    """
    if highlight_cells is None:
        highlight_cells = {}

    wb = openpyxl.load_workbook(str(file_path), data_only=True)
    sheets = []

    for ws in wb.worksheets:
        sheet_name = ws.title
        hl_set = set(highlight_cells.get(sheet_name, []))

        last_row, max_col = _get_data_bounds(ws)

        # 結合セルの情報をまとめる
        merged = {}  # {(min_row, min_col): (rowspan, colspan)}
        merged_skip = set()  # 結合で隠れるセル座標
        for mr in ws.merged_cells.ranges:
            r1, c1 = mr.min_row, mr.min_col
            rowspan = mr.max_row - mr.min_row + 1
            colspan = mr.max_col - mr.min_col + 1
            merged[(r1, c1)] = (rowspan, colspan)
            for rr in range(r1, mr.max_row + 1):
                for cc in range(c1, mr.max_col + 1):
                    if (rr, cc) != (r1, c1):
                        merged_skip.add((rr, cc))

        # 列幅 → colgroup
        colgroup = '<colgroup>'
        for c in range(1, max_col + 1):
            col_letter = get_column_letter(c)
            dim = ws.column_dimensions.get(col_letter)
            if dim and dim.width:
                # Excel幅 ≒ 文字数、1文字 ≈ 7px
                px = max(int(dim.width * 7), 20)
                colgroup += f'<col style="width:{px}px">'
            else:
                colgroup += '<col style="width:56px">'
        colgroup += '</colgroup>'

        # テーブル組み立て
        html_parts = [f'<table class="excel-sheet">{colgroup}<tbody>']

        for r in range(1, last_row + 1):
            html_parts.append("<tr>")
            for c in range(1, max_col + 1):
                if (r, c) in merged_skip:
                    continue

                cell = ws.cell(row=r, column=c)
                cell_ref = f"{get_column_letter(c)}{r}"

                # CSSスタイル
                style = _cell_style_css(cell)

                # ハイライト判定
                is_highlight = cell_ref in hl_set
                css_class = ' class="highlight"' if is_highlight else ''

                # 結合属性
                span_attrs = ""
                if (r, c) in merged:
                    rs, cs = merged[(r, c)]
                    if rs > 1:
                        span_attrs += f' rowspan="{rs}"'
                    if cs > 1:
                        span_attrs += f' colspan="{cs}"'

                style_attr = f' style="{style}"' if style else ''
                value = _format_cell_value(cell)

                html_parts.append(
                    f"<td{css_class}{span_attrs}{style_attr}>{value}</td>"
                )
            html_parts.append("</tr>")

        html_parts.append("</tbody></table>")

        sheets.append({
            "name": sheet_name,
            "html": "\n".join(html_parts),
        })

    wb.close()
    return sheets
